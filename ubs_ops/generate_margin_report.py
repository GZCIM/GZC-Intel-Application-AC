import csv
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime

def parse_number(value):
    """Parse a string number, handling empty strings and spaces"""
    if not value or value.strip() == '':
        return Decimal('0')
    try:
        return Decimal(str(value).strip().replace(',', ''))
    except:
        return Decimal('0')

def calculate_margin_summary(csv_file):
    """Calculate margin summary from CSV data"""

    # Initialize totals
    long_positions_mv = Decimal('0')
    short_positions_mv = Decimal('0')
    otc_mtm_mv = Decimal('0')
    money_market_mv = Decimal('0')
    net_cash_mv = Decimal('0')

    cross_margined_req = Decimal('0')
    otc_cross_netted_req = Decimal('0')
    money_market_margin_req = Decimal('0')
    long_short_benefit = Decimal('0')

    account_number = None
    account_name = None

    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)

        for row in reader:
            if account_number is None:
                account_number = row.get('Account', '').strip()
                account_name = row.get('Account_Name', '').strip()

            margin_type = row.get('Margin_Type', '').strip()
            product = row.get('Product', '').strip()
            reporting_group = row.get('Reporting_Group', '').strip()

            mv_rollup = parse_number(row.get('MV_Rollup', '0'))
            margin_rollup = parse_number(row.get('Margin_Rollup', '0'))

            # Money Market Funds
            if margin_type == 'CrossMarginPosition' and product == 'MMS':
                money_market_mv += mv_rollup
                money_market_margin_req += margin_rollup

            # Cross-Margined Requirement (from CrossMarginPosition)
            elif margin_type == 'CrossMarginPosition':
                cross_margined_req += margin_rollup

            # OTC MTM (Forward, Option, Swap - but not FX Margin Requirement)
            elif margin_type == 'CrossNetOTC':
                if reporting_group in ['Forward', 'Option', 'Swap']:
                    otc_mtm_mv += mv_rollup
                # OTC Cross-Netted Requirement (all OTC margin requirements)
                otc_cross_netted_req += margin_rollup

            # Net Cash (Cash Balances)
            elif margin_type == 'Cash Balances':
                net_cash_mv += mv_rollup

    # Calculate totals
    total_market_value = long_positions_mv + short_positions_mv + otc_mtm_mv + money_market_mv + net_cash_mv
    total_margin = cross_margined_req + otc_cross_netted_req + money_market_margin_req + long_short_benefit
    excess = total_market_value - total_margin

    # Round to 2 decimal places
    def round_decimal(d):
        return d.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    return {
        'account_number': account_number,
        'account_name': account_name,
        'summary': {
            'Long Positions': {
                'Market Value': round_decimal(long_positions_mv),
                'Margin': None
            },
            'Short Positions': {
                'Market Value': round_decimal(short_positions_mv),
                'Margin': None
            },
            'OTC MTM': {
                'Market Value': round_decimal(otc_mtm_mv),
                'Margin': None
            },
            'Money Market Funds': {
                'Market Value': round_decimal(money_market_mv),
                'Margin': None
            },
            'Net Cash': {
                'Market Value': round_decimal(net_cash_mv),
                'Margin': None
            },
            'Cross-Margined Requirement': {
                'Market Value': None,
                'Margin': round_decimal(cross_margined_req)
            },
            'OTC Cross-Netted Requirement': {
                'Market Value': None,
                'Margin': round_decimal(otc_cross_netted_req)
            },
            'Money Market Funds Margin Requirement': {
                'Market Value': None,
                'Margin': round_decimal(money_market_margin_req)
            },
            'Long Short Benefit': {
                'Market Value': None,
                'Margin': round_decimal(long_short_benefit)
            },
            'TOTAL': {
                'Market Value': round_decimal(total_market_value),
                'Margin': round_decimal(total_margin)
            },
            'Excess': {
                'Market Value': None,
                'Margin': round_decimal(excess)
            }
        }
    }

def format_currency(value):
    """Format currency value, handling None"""
    if value is None:
        return ''
    return f"{value:,.2f}"

def generate_html_report(data, output_file='margin_summary_report.html'):
    """Generate HTML report similar to PDF page 4"""

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Margin Summary Report</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 40px;
            background-color: #f5f5f5;
        }}
        .container {{
            background-color: white;
            padding: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .header {{
            margin-bottom: 30px;
        }}
        .header h1 {{
            color: #333;
            margin: 0;
            font-size: 24px;
        }}
        .account-info {{
            margin-top: 10px;
            font-size: 14px;
            color: #666;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background-color: #4a90e2;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr.total-row {{
            background-color: #e8f4f8;
            font-weight: bold;
            border-top: 2px solid #4a90e2;
        }}
        tr.excess-row {{
            background-color: #d4edda;
            font-weight: bold;
        }}
        .number {{
            text-align: right;
            font-family: 'Courier New', monospace;
        }}
        .footer {{
            margin-top: 30px;
            font-size: 12px;
            color: #999;
            text-align: center;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Margin Summary In USD</h1>
            <div class="account-info">
                <strong>Account Number:</strong> {data['account_number']}<br>
                <strong>Account Name:</strong> {data['account_name']}<br>
                <strong>Report Date:</strong> {datetime.now().strftime('%m/%d/%Y')}
            </div>
        </div>

        <table>
            <thead>
                <tr>
                    <th>Margin Summary In USD</th>
                    <th class="number">Market Value</th>
                    <th class="number">Margin</th>
                </tr>
            </thead>
            <tbody>
"""

    for category, values in data['summary'].items():
        row_class = ''
        if category == 'TOTAL':
            row_class = 'total-row'
        elif category == 'Excess':
            row_class = 'excess-row'

        mv_str = format_currency(values['Market Value'])
        margin_str = format_currency(values['Margin'])

        html += f"""                <tr class="{row_class}">
                    <td>{category}</td>
                    <td class="number">{mv_str}</td>
                    <td class="number">{margin_str}</td>
                </tr>
"""

    html += """            </tbody>
        </table>

        <div class="footer">
            Generated from CSV data - Margin Summary Report
        </div>
    </div>
</body>
</html>
"""

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"HTML report generated: {output_file}")

def export_to_excel(data, output_file='margin_summary_report.xlsx'):
    """Export to Excel format using openpyxl if available"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Margin Summary"

        # Header
        ws['A1'] = 'Margin Summary In USD'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Account Number: {data['account_number']}"
        ws['A3'] = f"Account Name: {data['account_name']}"
        ws['A4'] = f"Report Date: {datetime.now().strftime('%m/%d/%Y')}"

        # Table headers
        headers = ['Margin Summary In USD', 'Market Value', 'Margin']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')

        # Data rows
        row_num = 7
        for category, values in data['summary'].items():
            ws.cell(row=row_num, column=1, value=category)

            mv = values['Market Value']
            margin = values['Margin']

            if mv is not None:
                ws.cell(row=row_num, column=2, value=float(mv))
                ws.cell(row=row_num, column=2).number_format = '#,##0.00'

            if margin is not None:
                ws.cell(row=row_num, column=3, value=float(margin))
                ws.cell(row=row_num, column=3).number_format = '#,##0.00'

            # Style TOTAL and Excess rows
            if category == 'TOTAL':
                for col in range(1, 4):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
            elif category == 'Excess':
                for col in range(1, 4):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

            row_num += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18

        wb.save(output_file)
        print(f"Excel report generated: {output_file}")
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
        print("Falling back to CSV export...")
        export_to_csv(data, output_file.replace('.xlsx', '.csv'))

def export_to_csv(data, output_file='margin_summary_report.csv'):
    """Export margin summary to CSV"""
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Account Number', data['account_number']])
        writer.writerow(['Account Name', data['account_name']])
        writer.writerow(['Report Date', datetime.now().strftime('%m/%d/%Y')])
        writer.writerow([])
        writer.writerow(['Margin Summary In USD', 'Market Value', 'Margin'])

        for category, values in data['summary'].items():
            mv = format_currency(values['Market Value'])
            margin = format_currency(values['Margin'])
            writer.writerow([category, mv, margin])

    print(f"CSV report generated: {output_file}")

if __name__ == '__main__':
    csv_file = r'c:\tmp\20251110.MFXCMDRCSV.I0004255.CSV'

    print("Generating Margin Summary Report from CSV data...")
    result = calculate_margin_summary(csv_file)

    # Generate reports
    generate_html_report(result)
    export_to_excel(result)
    export_to_csv(result)

    print("\n" + "="*80)
    print("Report generation complete!")
    print("="*80)

